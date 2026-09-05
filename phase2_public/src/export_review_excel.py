"""Export gated review JSON/JSONL to a human-review Excel workbook.

The exporter does not infer legal conclusions. It flattens already-gated model
output, preserves long text, adds review columns and forces the overall status
to human second review. No input file is modified.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "Issue ID",
    "Test result",
    "Contract original",
    "Conclusion",
    "Risk category",
    "Legal basis",
    "Evidence boundary",
    "Assistant recommendation",
    "Processing suggestion",
    "Source locator",
    "Overall status",
    "Human reviewer decision",
    "Human reviewer notes",
]


def read_records(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".jsonl":
        return [
            json.loads(line)
            for line in path.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ]
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list):
        return payload
    if not isinstance(payload, dict):
        raise TypeError("Input JSON must be an object, list, or JSONL records")
    for key in ("results", "cases", "records"):
        if isinstance(payload.get(key), list):
            return payload[key]
    return [payload]


def gated_response(record: dict[str, Any]) -> tuple[dict[str, Any], str, bool]:
    gate = record.get("stage3_gate_result") or record.get("gate_result") or {}
    if not isinstance(gate, dict):
        gate = {}
    response = gate.get("response") or record.get("response") or record.get("output") or record
    if not isinstance(response, dict):
        response = {}
    status = str(gate.get("status") or record.get("gate_status") or record.get("status") or "")
    blocked = bool(gate.get("blocked") or record.get("blocked") or status == "blocked")
    return response, status or ("blocked" if blocked else "review_required"), blocked


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, (int, float, bool)):
        return str(value)
    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def legal_basis_text(items: Any) -> str:
    if not isinstance(items, list):
        return as_text(items)
    lines: list[str] = []
    for item in items:
        if not isinstance(item, dict):
            lines.append(as_text(item))
            continue
        label = " ".join(
            part
            for part in (
                as_text(item.get("law") or item.get("title")),
                as_text(item.get("article")),
                as_text(item.get("source_locator")),
            )
            if part
        )
        chunk_id = as_text(item.get("chunk_id"))
        if chunk_id:
            label = f"{label} [chunk_id={chunk_id}]".strip()
        lines.append(label)
    return "\n".join(line for line in lines if line)


def source_locator_text(finding: dict[str, Any]) -> str:
    location = finding.get("document_location")
    if isinstance(location, dict):
        preferred = [
            location.get("page"),
            location.get("section"),
            location.get("clause"),
            location.get("paragraph"),
            location.get("coordinates"),
        ]
        text = " | ".join(as_text(value) for value in preferred if value not in (None, ""))
        return text or as_text(location)
    return as_text(location or finding.get("source_locator"))


def flatten_record(record: dict[str, Any]) -> Iterable[list[str]]:
    response, gate_status, blocked = gated_response(record)
    issue_id = as_text(record.get("issue_id") or response.get("issue_id"))
    findings = response.get("findings")
    if not isinstance(findings, list) or not findings:
        yield [
            issue_id,
            "blocked" if blocked else gate_status,
            "",
            "blocked" if blocked else as_text(response.get("conclusion_type")),
            "",
            "",
            "Core response was blocked or contained no deliverable finding.",
            "Send the source material and gate diagnostics for human review.",
            "rejected" if blocked else "revised",
            "",
            "requires_human_second_review",
            "",
            "",
        ]
        return

    for finding in findings:
        if not isinstance(finding, dict):
            continue
        recommendation = finding.get("assistant_recommendation")
        if isinstance(recommendation, dict):
            recommendation = (
                recommendation.get("substantive_conclusion")
                or recommendation.get("recommended_handling")
                or recommendation
            )
        yield [
            as_text(finding.get("issue_id") or issue_id),
            gate_status,
            as_text(finding.get("document_excerpt")),
            as_text(finding.get("conclusion_type") or finding.get("reasoning_conclusion")),
            as_text(finding.get("risk_category")),
            legal_basis_text(finding.get("legal_evidence")),
            as_text(finding.get("evidence_boundary")),
            as_text(recommendation or finding.get("recommended_human_action")),
            as_text(finding.get("review_processing_label") or "revised"),
            source_locator_text(finding),
            "requires_human_second_review",
            "",
            "",
        ]


def build_workbook(records: Iterable[dict[str, Any]], output_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Human review"
    sheet.append(HEADERS)

    for record in records:
        for row in flatten_record(record):
            sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    alert_fill = PatternFill("solid", fgColor="F4CCCC")
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        status = str(row[1].value or "").lower()
        conclusion = str(row[3].value or "").lower()
        if "blocked" in status or "legal_confirm" in conclusion:
            for cell in row:
                cell.fill = alert_fill
        elif "review" in status or "review" in conclusion or "insufficient" in conclusion:
            for cell in row:
                cell.fill = review_fill

    widths = [18, 14, 52, 38, 24, 60, 52, 60, 20, 30, 30, 24, 48]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 32

    info = workbook.create_sheet("Read me")
    info.append(["Field", "Meaning"])
    info.append(["Purpose", "Human review aid; not a final legal, award, rejection, or disqualification decision."])
    info.append(["Processing suggestion", "Assistant-proposed accepted/revised/rejected status; a human reviewer must decide."])
    info.append(["Overall status", "Always requires_human_second_review for generated rows."])
    info.append(["Input integrity", "The exporter only flattens gated JSON/JSONL and does not modify the source file."])
    for cell in info[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 100
    for row in info.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    build_workbook(read_records(args.input), args.output)
    print(json.dumps({"output": str(args.output), "status": "written"}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
