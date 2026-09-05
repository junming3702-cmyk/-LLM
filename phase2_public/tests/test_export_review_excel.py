from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from export_review_excel import HEADERS, build_workbook


class ExportReviewExcelTests(unittest.TestCase):
    def test_gated_finding_is_exported_without_formula_cells(self) -> None:
        record = {
            "issue_id": "SYN-TEST-001",
            "stage3_gate_result": {
                "status": "corrected",
                "blocked": False,
                "response": {
                    "findings": [
                        {
                            "issue_id": "SYN-TEST-001",
                            "document_excerpt": "Clause 1 requires a three-day period.",
                            "conclusion_type": "requires_human_legal_review",
                            "risk_category": "time requirement",
                            "legal_evidence": [
                                {
                                    "law": "Synthetic regulation",
                                    "article": "Article 1",
                                    "source_locator": "page 1",
                                    "chunk_id": "chunk-001",
                                }
                            ],
                            "evidence_boundary": "Synthetic fixture only.",
                            "assistant_recommendation": "Compare the dates and confirm applicability.",
                            "review_processing_label": "revised",
                            "document_location": {"page": 3, "clause": "1"},
                        }
                    ]
                },
            },
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "review.xlsx"
            build_workbook([record], output)
            workbook = load_workbook(output, data_only=False)
            sheet = workbook["Human review"]
            self.assertEqual([cell.value for cell in sheet[1]], HEADERS)
            self.assertEqual(sheet["A2"].value, "SYN-TEST-001")
            self.assertEqual(sheet["K2"].value, "requires_human_second_review")
            self.assertIn("Synthetic regulation", sheet["F2"].value)
            self.assertFalse(
                any(
                    isinstance(cell.value, str) and cell.value.startswith("=")
                    for row in sheet.iter_rows()
                    for cell in row
                )
            )


if __name__ == "__main__":
    unittest.main()
